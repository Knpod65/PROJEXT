from __future__ import annotations

import io
import os
import sys
from datetime import date
from types import SimpleNamespace

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import models
from auth_utils import get_current_user
from database import Base, get_db
from routers.invigilation_advance_batch import router
from services.payment_supporting_finance_roster_service import (
    DRAFT_LABEL_EN,
    EXPORT_STATUS,
    SHEET_NAMES,
    _build_data,
    build_finance_support_roster_export,
    load_supporting_roster_sources,
)
from services.thai_export_service import XLSX_THAI_FONT, has_mojibake_marker

EXPORT_URL = "/api/invigilation-advance-batch/finance-support-roster-export"
STATUS_URL = "/api/invigilation-advance-batch/finance-support-roster-status"
REQUEST = {"academic_year": "2568", "semester": "2", "exam_type": "final", "period_id": 1}


@pytest.fixture()
def app_and_db():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(bind=engine)
    TestingSession = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    session = TestingSession()
    app = FastAPI()
    app.include_router(router, prefix="/api/invigilation-advance-batch")
    app.dependency_overrides[get_db] = lambda: session
    yield app, session
    session.close()


def _client(app, role=models.UserRole.admin):
    app.dependency_overrides[get_current_user] = lambda: SimpleNamespace(
        id=999, role=role, view_as_role=None, _active_role=role
    )
    return TestClient(app)


def _seed(session, *, room_counts=(70, 40), online=False, same_supervisor=False, paper_users=2):
    period = models.ExamPeriod(id=1, academic_year="2568", semester="2", exam_type="final", label="ปลายภาค 2/2568")
    users = [
        models.User(id=1, username="chief", email="chief@example.test", password_hash="x", role=models.UserRole.staff, full_name="หัวหน้าคุมสอบ"),
        models.User(id=2, username="keeper", email="keeper@example.test", password_hash="x", role=models.UserRole.staff, full_name="ผู้เปิดห้อง"),
        models.User(id=3, username="paper1", email="paper1@example.test", password_hash="x", role=models.UserRole.staff, full_name="ผู้จ่ายข้อสอบหนึ่ง"),
        models.User(id=4, username="paper2", email="paper2@example.test", password_hash="x", role=models.UserRole.staff, full_name="ผู้จ่ายข้อสอบสอง"),
        models.User(id=5, username="paper3", email="paper3@example.test", password_hash="x", role=models.UserRole.staff, full_name="ผู้จ่ายข้อสอบสาม"),
    ]
    session.add_all([period, *users])
    for index, count in enumerate(room_counts, 1):
        room = models.Room(id=index, room_name=f"R{index:02d}", capacity=100, e_room_code=f"E{index}" if online else None)
        course = models.Course(id=index, course_id=f"10010{index}", course_name_th=f"วิชา {index}")
        section = models.Section(id=index, course_id=index, section_no=str(index), teacher_id=1, num_students=count, semester="2", academic_year="2568")
        schedule = models.ExamSchedule(
            id=index, section_id=index, room_id=index, exam_date=date(2026, 6, 20),
            exam_time="09.00-12.00", exam_time_start="09:00", exam_time_end="12:00",
            exam_type=models.ExamType.final, status=models.ScheduleStatus.published,
            paper_distributor="LEGACY MUST NOT BE USED",
        )
        user_id = 1 if same_supervisor or index == 1 else 2
        role = "chief" if user_id == 1 else "room_keeper"
        session.add_all([room, course, section, schedule, models.Supervision(schedule_id=index, user_id=user_id, role_in_exam=role)])
    for index in range(paper_users):
        session.add(models.PaperDistributionAssignment(
            exam_period_id=1, user_id=3 + index, exam_date="2026-06-20", exam_time="09.00-12.00",
            start_time="09:00", end_time="12:00", slot_order=index + 1,
            duty_type=models.StaffDutyType.paper_distribution,
        ))
    session.add(models.PaymentDocumentSettings(
        settings_id="payment-document-settings-2-2568", term="2/2568", weekday_rate=120, weekend_rate=200,
        currency="THB", payment_unit="PER_PERSON_SESSION", paper_distribution_responsible_group="ESQ",
        status=models.PaymentDocumentSettingsStatus.active_for_draft_preview,
        payment_authorization_enabled=False, final_export_enabled=False,
    ))
    session.add(models.PaymentDocumentReviewRecord(
        document_id="ADVANCE_PAYMENT_DRAFT_SUMMARY:2568:2:final:1", document_type="ADVANCE_PAYMENT_DRAFT_SUMMARY",
        term="2/2568", review_status="ACCEPTED_FOR_DRAFT_EXPORT", comment="Accepted for draft supporting export",
        reviewer_name="Reviewer", reviewer_role="admin", payment_authorization_enabled=False, final_export_enabled=False,
    ))
    session.commit()


def _workbook(response):
    return openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)


def test_live_supervision_room_keeper_pda_and_legacy_excluded(app_and_db):
    app, session = app_and_db
    _seed(session)
    data = _build_data(load_supporting_roster_sources(session, REQUEST))
    assert {row["primary_role"] for row in data["people"]} >= {"chief", "room_keeper", "paper_distribution"}
    assert all("SupervisionBaseline" not in row["sources_text"] for row in data["people"])
    assert all("LEGACY" not in row["sources_text"] for row in data["people"])
    assert any("PaperDistributionAssignment" in row["sources_text"] for row in data["people"])


def test_same_person_multiple_rooms_same_slot_counts_once(app_and_db):
    _, session = app_and_db
    _seed(session, same_supervisor=True)
    data = _build_data(load_supporting_roster_sources(session, REQUEST))
    chief = [row for row in data["people"] if row["person_id"] == 1]
    assert len(chief) == 1
    assert chief[0]["payment_count"] == 1
    assert chief[0]["rooms_text"] == "R01, R02"
    assert "DUPLICATE_PERSON_COLLAPSED_TO_ONE_PAYMENT_COUNT" in chief[0]["status_text"]


def test_cross_source_duplicate_is_supervision_primary(app_and_db):
    _, session = app_and_db
    _seed(session, paper_users=0)
    session.add(models.PaperDistributionAssignment(
        exam_period_id=1, user_id=1, exam_date="2026-06-20", exam_time="09.00-12.00",
        start_time="09:00", end_time="12:00", slot_order=1, duty_type=models.StaffDutyType.paper_distribution,
    ))
    session.commit()
    data = _build_data(load_supporting_roster_sources(session, REQUEST))
    row = next(row for row in data["people"] if row["person_id"] == 1)
    assert row["category"] == "invigilation"
    assert row["primary_role"] == "chief"
    assert "PaperDistributionAssignment" in row["sources_text"]
    assert data["summaries"][0]["paper_count"] == 0


def test_top_two_room_mapping_and_tie_break_are_deterministic(app_and_db):
    _, session = app_and_db
    _seed(session, room_counts=(40, 70))
    data = _build_data(load_supporting_roster_sources(session, REQUEST))
    mapped = [row for row in data["mappings"] if row["status"] == "MAPPED_TO_TOP_ROOM"]
    assert [(row["rank"], row["room"], row["person"]) for row in mapped] == [
        (1, "R02", "ผู้จ่ายข้อสอบหนึ่ง"), (2, "R01", "ผู้จ่ายข้อสอบสอง")
    ]


def test_single_room_and_extra_person_flags(app_and_db):
    _, session = app_and_db
    _seed(session, room_counts=(70,), paper_users=2)
    data = _build_data(load_supporting_roster_sources(session, REQUEST))
    assert data["mappings"][0]["status"] == "MAPPED_TO_ONLY_ELIGIBLE_ROOM"
    assert data["mappings"][1]["status"] == "EXTRA_PAPER_DISTRIBUTION_REVIEW_REQUIRED"
    assert data["summaries"][0]["paper_count"] == 1


def test_all_online_is_trace_only_and_not_paper_payable(app_and_db):
    _, session = app_and_db
    _seed(session, online=True)
    data = _build_data(load_supporting_roster_sources(session, REQUEST))
    assert all(row["note"] == "TRACE_ONLY_ONLINE_ROW" for row in data["traces"])
    assert all(row["status"] == "NO_ELIGIBLE_PHYSICAL_ROOM_FOR_PAPER_MAPPING" for row in data["mappings"])
    assert data["summaries"][0]["paper_count"] == 0


def test_missing_paper_assignment_is_flagged(app_and_db):
    _, session = app_and_db
    _seed(session, paper_users=1)
    data = _build_data(load_supporting_roster_sources(session, REQUEST))
    assert any(row["status"] == "MISSING_PAPER_DISTRIBUTION_ASSIGNMENT" for row in data["mappings"])


def test_export_has_exact_five_sheets_warnings_and_safety_metadata(app_and_db):
    app, session = app_and_db
    _seed(session, same_supervisor=True)
    response = _client(app).post(EXPORT_URL, json=REQUEST)
    assert response.status_code == 200
    assert "filename*=UTF-8''" in response.headers.get("content-disposition", "")
    workbook = _workbook(response)
    assert workbook.sheetnames == SHEET_NAMES
    assert workbook.sheetnames[0] == "สรุปตามวันและช่วงเวลา"
    assert workbook.sheetnames[2] == "ใบลงลายมือชื่อประกอบการเบิก"
    for sheet in workbook.worksheets:
        assert sheet["A2"].value == DRAFT_LABEL_EN
        assert sheet["A3"].value == "DRAFT_NOT_AUTHORIZED"
        assert sheet["A1"].font.name == XLSX_THAI_FONT
        assert sheet["A5"].font.name == XLSX_THAI_FONT
        assert not has_mojibake_marker(sheet.title)
        for row in sheet.iter_rows():
            for cell in row:
                assert not has_mojibake_marker(cell.value)
    roster = workbook[SHEET_NAMES[1]]
    values = [roster.cell(row, 2).value for row in range(1, roster.max_row + 1)]
    assert EXPORT_STATUS in values
    assert "false" in values


def test_status_and_export_gate_missing_settings(app_and_db):
    app, session = app_and_db
    _seed(session)
    session.query(models.PaymentDocumentSettings).delete()
    session.commit()
    status = _client(app).post(STATUS_URL, json=REQUEST)
    export = _client(app).post(EXPORT_URL, json=REQUEST)
    assert status.status_code == 200
    assert status.json()["available"] is False
    assert export.status_code == 400


def test_requested_period_must_match_scope(app_and_db):
    app, session = app_and_db
    _seed(session)
    bad_request = {**REQUEST, "period_id": 999}
    status = _client(app).post(STATUS_URL, json=bad_request)
    assert status.status_code == 200
    assert status.json()["available"] is False
    assert "Requested exam period" in "; ".join(status.json()["blocked_reasons"])


@pytest.mark.parametrize("role", [models.UserRole.staff, models.UserRole.teacher, models.UserRole.print_shop, models.UserRole.student])
def test_unauthorized_roles_are_blocked(app_and_db, role):
    app, session = app_and_db
    _seed(session)
    assert _client(app, role).post(EXPORT_URL, json=REQUEST).status_code == 403
    assert _client(app, role).post(STATUS_URL, json=REQUEST).status_code == 403


@pytest.mark.parametrize("role", [models.UserRole.admin, models.UserRole.esq_head, models.UserRole.secretary])
def test_reviewer_roles_are_allowed(app_and_db, role):
    app, session = app_and_db
    _seed(session)
    assert _client(app, role).post(EXPORT_URL, json=REQUEST).status_code == 200


def test_export_is_read_only_and_rc1_route_still_exists(app_and_db):
    app, session = app_and_db
    _seed(session)
    counts_before = {
        "supervision": session.query(models.Supervision).count(),
        "paper": session.query(models.PaperDistributionAssignment).count(),
        "review": session.query(models.PaymentDocumentReviewRecord).count(),
    }
    assert _client(app).post(EXPORT_URL, json=REQUEST).status_code == 200
    counts_after = {
        "supervision": session.query(models.Supervision).count(),
        "paper": session.query(models.PaperDistributionAssignment).count(),
        "review": session.query(models.PaymentDocumentReviewRecord).count(),
    }
    assert counts_after == counts_before
    assert any(getattr(route, "path", "") == "/official-document-draft-export" for route in router.routes)


def test_direct_builder_uses_safe_draft_filename(app_and_db):
    _, session = app_and_db
    _seed(session)
    workbook, filename = build_finance_support_roster_export(session, REQUEST)
    assert workbook.sheetnames == SHEET_NAMES
    assert filename.startswith("EMS_DRAFT_FINANCE_ROSTER_2-2568_")
    assert "FINAL" not in filename
