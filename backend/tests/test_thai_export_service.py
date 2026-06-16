from __future__ import annotations

import io

import openpyxl

from services.thai_export_service import (
    XLSX_THAI_FONT,
    apply_workbook_thai_style,
    content_disposition,
    csv_bytes,
    has_mojibake_marker,
    register_reportlab_thai_fonts,
    safe_ascii_filename,
)


def test_content_disposition_has_ascii_fallback_and_utf8_filename_star():
    header = content_disposition(
        "ร่างรายชื่อประกอบการเบิกค่าคุมสอบ_2-2568.xlsx",
        ascii_fallback="EMS_DRAFT_FINANCE_ROSTER_2-2568.xlsx",
    )

    assert 'filename="EMS_DRAFT_FINANCE_ROSTER_2-2568.xlsx"' in header
    assert "filename*=UTF-8''" in header
    assert "%E0%B8%A3" in header
    assert "/" not in header.split("filename=", 1)[1].split(";", 1)[0]


def test_safe_ascii_filename_removes_slashes_controls_and_thai_fallback():
    assert safe_ascii_filename("../ร่าง/บัญชี\n.xlsx", default="EMS_EXPORT") == "EMS_EXPORT.xlsx"
    assert safe_ascii_filename("EMS export: รอบ 2.xlsx") == "EMS_export_2.xlsx"


def test_csv_bytes_are_utf8_sig_and_round_trip_thai_text():
    data = csv_bytes(
        ["ชื่อ", "จำนวนเงิน"],
        [["คณะรัฐศาสตร์ฯ", "1,200.00"], ["ข้อความยาว – พร้อมเลข 123", "0"]],
    )

    assert data.startswith(b"\xef\xbb\xbf")
    decoded = data.decode("utf-8-sig")
    assert "คณะรัฐศาสตร์ฯ" in decoded
    assert "ข้อความยาว – พร้อมเลข 123" in decoded
    assert not has_mojibake_marker(decoded)


def test_workbook_thai_style_round_trips_values_and_font_names():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "รายงานภาษาไทย"
    sheet["A1"] = "ชื่อ"
    sheet["A2"] = "คณะรัฐศาสตร์ฯ"
    sheet["B2"] = "ร่างเอกสารเพื่อการตรวจทานเท่านั้น"

    apply_workbook_thai_style(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    reopened = openpyxl.load_workbook(buffer)

    assert reopened.sheetnames == ["รายงานภาษาไทย"]
    assert reopened.active["A2"].value == "คณะรัฐศาสตร์ฯ"
    assert reopened.active["A1"].font.name == XLSX_THAI_FONT
    assert reopened.active["B2"].font.name == XLSX_THAI_FONT
    assert reopened.active["B2"].alignment.wrap_text is True


def test_reportlab_font_registration_contract_is_explicit():
    registration = register_reportlab_thai_fonts("EMS-Test-Thai")

    assert registration.normal_name
    assert registration.bold_name
    assert isinstance(registration.embedded, bool)
    if registration.embedded:
        assert registration.normal_path
        assert registration.bold_path
    else:
        assert registration.normal_name == "Helvetica"
