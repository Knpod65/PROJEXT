"""Draft-only supporting finance roster XLSX export.

Uses live Supervision and PaperDistributionAssignment rows. The service is
read-only and cannot authorize payment or create an official final export.
"""
from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any

from fastapi import HTTPException
from sqlalchemy.orm import Session, joinedload

import models
from services.official_payment_document_draft_service import (
    _normalize_exam_date,
    _normalize_slot,
    _term_context,
)
from services.payment_document_draft_export_service import (
    ACTIVE_SETTINGS_STATUS,
    _document_id,
    _find_accepted_review,
)
from services.payment_document_settings_service import (
    SETTINGS_CONFIGURED,
    resolve_payment_document_settings_for_draft,
)
from services.thai_export_service import apply_workbook_thai_style

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


DRAFT_LABEL_TH = "ร่างเอกสารประกอบการตรวจนับค่าตอบแทน ยังไม่ใช่เอกสารอนุมัติเบิกจ่าย"
DRAFT_LABEL_EN = "Draft supporting roster for payment checking only. Not payment authorization."
DRAFT_STATUS = "DRAFT_NOT_AUTHORIZED"
EXPORT_STATUS = "DRAFT_SUPPORTING_EXPORT_ONLY"
ROLE_PRECEDENCE = {"chief": 0, "room_keeper": 1, "supervisor": 2, "distributor": 3, "paper_distribution": 4}
ROLE_LABELS = {
    "chief": "กรรมการคุมสอบ (หัวหน้า)",
    "room_keeper": "ผู้เปิดห้อง/คุมสอบ",
    "supervisor": "กรรมการคุมสอบ",
    "distributor": "กรรมการคุมสอบ (distributor)",
    "paper_distribution": "กรรมการจ่ายข้อสอบ/เปิดห้อง",
}
SHEET_NAMES = [
    "สรุปตามวันและช่วงเวลา",
    "รายชื่อประกอบการเบิก",
    "ใบลงลายมือชื่อประกอบการเบิก",
    "การผูกคนจ่ายข้อสอบกับห้อง",
    "รายละเอียดอ้างอิงห้องและวิชา",
]


def _value(value: Any) -> Any:
    return value.value if hasattr(value, "value") else value


def _scope_payload(payload: dict[str, Any]) -> dict[str, Any]:
    result = dict(payload)
    for key in ("date_from", "date_to"):
        value = result.get(key)
        if isinstance(value, date):
            result[key] = value.isoformat()
    if result.get("date_from") and result.get("date_to") and result["date_to"] < result["date_from"]:
        raise HTTPException(400, "date_to must not be earlier than date_from")
    return result


def _periods_for_scope(db: Session, payload: dict[str, Any]) -> list[models.ExamPeriod]:
    query = db.query(models.ExamPeriod)
    if payload.get("period_id") is not None:
        query = query.filter(models.ExamPeriod.id == payload["period_id"])
    if payload.get("academic_year"):
        query = query.filter(models.ExamPeriod.academic_year == str(payload["academic_year"]))
    if payload.get("semester"):
        query = query.filter(models.ExamPeriod.semester == str(payload["semester"]))
    if payload.get("exam_type"):
        query = query.filter(models.ExamPeriod.exam_type == str(payload["exam_type"]))
    return query.order_by(models.ExamPeriod.id).all()


def load_supporting_roster_sources(db: Session, request_payload: dict[str, Any]) -> dict[str, Any]:
    payload = _scope_payload(request_payload)
    periods = _periods_for_scope(db, payload)
    query = db.query(models.ExamSchedule).options(
        joinedload(models.ExamSchedule.section).joinedload(models.Section.course),
        joinedload(models.ExamSchedule.section).joinedload(models.Section.teacher),
        joinedload(models.ExamSchedule.room),
        joinedload(models.ExamSchedule.supervisions).joinedload(models.Supervision.user),
    ).join(models.Section).filter(
        models.ExamSchedule.status.in_([models.ScheduleStatus.published, models.ScheduleStatus.locked]),
    )
    if payload.get("academic_year"):
        query = query.filter(models.Section.academic_year == str(payload["academic_year"]))
    if payload.get("semester"):
        query = query.filter(models.Section.semester == str(payload["semester"]))
    if payload.get("exam_type"):
        query = query.filter(models.ExamSchedule.exam_type == str(payload["exam_type"]))
    if payload.get("date_from"):
        query = query.filter(models.ExamSchedule.exam_date >= payload["date_from"])
    if payload.get("date_to"):
        query = query.filter(models.ExamSchedule.exam_date <= payload["date_to"])
    schedules = query.order_by(models.ExamSchedule.exam_date, models.ExamSchedule.exam_time, models.ExamSchedule.id).all()

    period_ids = [period.id for period in periods]
    assignments: list[models.PaperDistributionAssignment] = []
    if period_ids:
        paper_query = db.query(models.PaperDistributionAssignment).options(
            joinedload(models.PaperDistributionAssignment.user)
        ).filter(models.PaperDistributionAssignment.exam_period_id.in_(period_ids))
        if payload.get("date_from"):
            paper_query = paper_query.filter(models.PaperDistributionAssignment.exam_date >= payload["date_from"])
        if payload.get("date_to"):
            paper_query = paper_query.filter(models.PaperDistributionAssignment.exam_date <= payload["date_to"])
        assignments = paper_query.order_by(
            models.PaperDistributionAssignment.exam_date,
            models.PaperDistributionAssignment.exam_time,
            models.PaperDistributionAssignment.slot_order,
            models.PaperDistributionAssignment.id,
        ).all()
    term = _term_context(payload, periods)
    settings = resolve_payment_document_settings_for_draft(db, term=term["term_label"] or None)
    return {"payload": payload, "periods": periods, "schedules": schedules, "paper_assignments": assignments, "term": term, "settings": settings}


def _slot_info(schedule: models.ExamSchedule, rates: dict[str, Decimal] | None) -> dict[str, Any]:
    date_info = _normalize_exam_date(schedule.exam_date, rates)
    start, end, time_slot = _normalize_slot(schedule.exam_time_start, schedule.exam_time_end, schedule.exam_time)
    key = (date_info["normalized_exam_date"] or date_info["exam_date"], time_slot)
    return {**date_info, "start_time": start, "end_time": end, "time_slot": time_slot, "key": key}


def _paper_slot_info(assignment: models.PaperDistributionAssignment, rates: dict[str, Decimal] | None) -> dict[str, Any]:
    date_info = _normalize_exam_date(assignment.exam_date, rates)
    start, end, time_slot = _normalize_slot(assignment.start_time, assignment.end_time, assignment.exam_time)
    key = (date_info["normalized_exam_date"] or date_info["exam_date"], time_slot)
    return {**date_info, "start_time": start, "end_time": end, "time_slot": time_slot, "key": key}


def _build_data(source: dict[str, Any]) -> dict[str, Any]:
    settings = source["settings"]
    rates = None
    if settings["settings_source_status"] == SETTINGS_CONFIGURED:
        rates = {"WEEKDAY": settings["weekday_rate"], "WEEKEND": settings["weekend_rate"]}
    slots: dict[tuple[str, str], dict[str, Any]] = {}
    for schedule in source["schedules"]:
        info = _slot_info(schedule, rates)
        slot = slots.setdefault(info["key"], {**info, "schedules": [], "paper_assignments": []})
        slot["schedules"].append(schedule)
    for assignment in source["paper_assignments"]:
        info = _paper_slot_info(assignment, rates)
        slot = slots.get(info["key"])
        if slot is not None:
            slot["paper_assignments"].append(assignment)

    person_rows: list[dict[str, Any]] = []
    mapping_rows: list[dict[str, Any]] = []
    trace_rows: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []

    for slot in sorted(slots.values(), key=lambda row: row["key"]):
        room_map: dict[int, dict[str, Any]] = {}
        all_refs: set[str] = set()
        total_students = 0
        for schedule in slot["schedules"]:
            section = schedule.section
            course = section.course if section else None
            room = schedule.room
            students = int(getattr(section, "num_students", 0) or 0)
            course_code = str(getattr(course, "course_id", "") or "")
            section_no = str(getattr(section, "section_no", "") or "")
            ref = f"{course_code}/{section_no}".strip("/")
            all_refs.add(ref)
            total_students += students
            sup_names = sorted({(sup.user.full_name or sup.user.username) for sup in schedule.supervisions if sup.user})
            trace_rows.append({
                "date": slot["exam_date"], "time": slot["time_slot"], "course": course_code,
                "section": section_no, "instructor": getattr(getattr(section, "teacher", None), "full_name", None) or "",
                "students": students, "room": getattr(room, "room_name", None) or "",
                "room_type": "online" if room and room.e_room_code is not None else "physical",
                "supervision_names": ", ".join(sup_names), "paper_names": "", "schedule_id": schedule.id,
                "note": "TRACE_ONLY_ONLINE_ROW" if room and room.e_room_code is not None else "",
            })
            if room and room.e_room_code is None:
                aggregate = room_map.setdefault(room.id, {"room": room.room_name, "students": 0, "refs": set()})
                aggregate["students"] += students
                aggregate["refs"].add(ref)

        ranked_rooms = sorted(
            room_map.values(),
            key=lambda row: (-row["students"], row["room"], sorted(row["refs"])[0] if row["refs"] else ""),
        )
        paper_assignments = sorted(slot["paper_assignments"], key=lambda row: (row.slot_order or 0, row.id or 0))
        mapped_paper: dict[int, dict[str, Any]] = {}
        required = min(2, len(ranked_rooms))
        for index, assignment in enumerate(paper_assignments):
            person_name = assignment.user.full_name or assignment.user.username if assignment.user else ""
            if index < required:
                room = ranked_rooms[index]
                status = "MAPPED_TO_ONLY_ELIGIBLE_ROOM" if len(ranked_rooms) == 1 else "MAPPED_TO_TOP_ROOM"
                mapped_paper[assignment.user_id] = {"room": room["room"], "status": status}
                mapping_rows.append({
                    "date": slot["exam_date"], "time": slot["time_slot"], "rank": index + 1, "room": room["room"],
                    "students": room["students"], "refs": ", ".join(sorted(room["refs"])), "person": person_name,
                    "reason": "only-eligible-room" if len(ranked_rooms) == 1 else f"top-{index + 1}-by-student-count",
                    "status": status, "note": "",
                })
            else:
                status = "NO_ELIGIBLE_PHYSICAL_ROOM_FOR_PAPER_MAPPING" if not ranked_rooms else "EXTRA_PAPER_DISTRIBUTION_REVIEW_REQUIRED"
                mapping_rows.append({
                    "date": slot["exam_date"], "time": slot["time_slot"], "rank": None, "room": "",
                    "students": None, "refs": "", "person": person_name, "reason": "not-mapped",
                    "status": status, "note": "Trace/review only; not included in payable paper headcount.",
                })
        if required > len(paper_assignments):
            for index in range(len(paper_assignments), required):
                room = ranked_rooms[index]
                mapping_rows.append({
                    "date": slot["exam_date"], "time": slot["time_slot"], "rank": index + 1, "room": room["room"],
                    "students": room["students"], "refs": ", ".join(sorted(room["refs"])), "person": "",
                    "reason": "required-room-has-no-paper-person", "status": "MISSING_PAPER_DISTRIBUTION_ASSIGNMENT", "note": "",
                })

        people: dict[int, dict[str, Any]] = {}
        for schedule in slot["schedules"]:
            section = schedule.section
            course = section.course if section else None
            room_name = getattr(schedule.room, "room_name", None) or ""
            ref = f"{getattr(course, 'course_id', '')}/{getattr(section, 'section_no', '')}".strip("/")
            for supervision in schedule.supervisions:
                if not supervision.user_id or not supervision.user:
                    continue
                role = str(supervision.role_in_exam or "supervisor")
                row = people.setdefault(supervision.user_id, {
                    "date": slot["exam_date"], "normalized_date": slot["normalized_exam_date"], "time": slot["time_slot"],
                    "day_type": slot["day_type"], "person_id": supervision.user_id,
                    "name": supervision.user.full_name or supervision.user.username, "roles": set(), "rooms": set(),
                    "refs": set(), "sources": set(), "statuses": set(), "rate": slot["rate_amount"],
                })
                if row["roles"]:
                    row["statuses"].add("DUPLICATE_PERSON_COLLAPSED_TO_ONE_PAYMENT_COUNT")
                row["roles"].add(role)
                row["rooms"].add(room_name)
                row["refs"].add(ref)
                row["sources"].add("Supervision")
        for assignment in paper_assignments:
            if assignment.user_id not in mapped_paper or not assignment.user:
                continue
            mapping = mapped_paper[assignment.user_id]
            row = people.setdefault(assignment.user_id, {
                "date": slot["exam_date"], "normalized_date": slot["normalized_exam_date"], "time": slot["time_slot"],
                "day_type": slot["day_type"], "person_id": assignment.user_id,
                "name": assignment.user.full_name or assignment.user.username, "roles": set(), "rooms": set(),
                "refs": set(), "sources": set(), "statuses": set(), "rate": slot["rate_amount"],
            })
            if row["roles"]:
                row["statuses"].add("DUPLICATE_PERSON_COLLAPSED_TO_ONE_PAYMENT_COUNT")
            row["roles"].add("paper_distribution")
            row["rooms"].add(mapping["room"])
            row["sources"].add("PaperDistributionAssignment")
            row["statuses"].add(mapping["status"])

        inv_count = 0
        paper_count = 0
        for row in people.values():
            roles = sorted(row["roles"], key=lambda role: ROLE_PRECEDENCE.get(role, 99))
            primary_role = roles[0]
            category = "paper" if primary_role == "paper_distribution" else "invigilation"
            inv_count += category == "invigilation"
            paper_count += category == "paper"
            person_rows.append({
                **row, "primary_role": primary_role, "role_label": ROLE_LABELS.get(primary_role, primary_role),
                "category": category, "rooms_text": ", ".join(sorted(filter(None, row["rooms"]))),
                "refs_text": ", ".join(sorted(filter(None, row["refs"]))), "sources_text": ", ".join(sorted(row["sources"])),
                "status_text": ", ".join(sorted(row["statuses"])), "payment_count": 1,
                "amount": row["rate"], "note": f"All source roles: {', '.join(roles)}",
            })
        rate = slot["rate_amount"]
        summaries.append({
            "date": slot["exam_date"], "normalized_date": slot["normalized_exam_date"], "time": slot["time_slot"],
            "day_type": slot["day_type"], "weekday": date.fromisoformat(slot["normalized_exam_date"]).strftime("%A") if slot["normalized_exam_date"] else "",
            "physical_rooms": len(ranked_rooms), "course_sections": len(all_refs), "students": total_students,
            "inv_count": inv_count, "paper_count": paper_count, "total_count": inv_count + paper_count, "rate": rate,
            "inv_amount": rate * inv_count if rate is not None else None,
            "paper_amount": rate * paper_count if rate is not None else None,
            "total_amount": rate * (inv_count + paper_count) if rate is not None else None,
            "note": "; ".join(sorted({row["status"] for row in mapping_rows if row["date"] == slot["exam_date"] and row["time"] == slot["time_slot"] and row["status"] not in {"MAPPED_TO_TOP_ROOM"}})),
        })
        paper_names = ", ".join(sorted({assignment.user.full_name or assignment.user.username for assignment in paper_assignments if assignment.user}))
        for trace in trace_rows:
            if trace["date"] == slot["exam_date"] and trace["time"] == slot["time_slot"]:
                trace["paper_names"] = paper_names
    return {"summaries": summaries, "people": person_rows, "mappings": mapping_rows, "traces": trace_rows}


def _gate_status(db: Session, source: dict[str, Any]) -> dict[str, Any]:
    blockers: list[str] = []
    settings = source["settings"]
    if source["payload"].get("period_id") is not None and not source["periods"]:
        blockers.append("Requested exam period does not exist or does not match the requested term/exam type.")
    try:
        review = _find_accepted_review(db, _document_id(source["payload"]))
        review_status = review.review_status
        if review.payment_authorization_enabled:
            blockers.append("Accepted review payment_authorization_enabled must be false.")
        if review.final_export_enabled:
            blockers.append("Accepted review final_export_enabled must be false.")
    except HTTPException as exc:
        review_status = "BLOCKED"
        blockers.append(str(exc.detail))
    if settings["settings_source_status"] != SETTINGS_CONFIGURED:
        blockers.extend(settings["settings_issues"])
    if settings.get("settings_status") != ACTIVE_SETTINGS_STATUS:
        blockers.append("settings_status must be ACTIVE_FOR_DRAFT_PREVIEW")
    settings_record = None
    if settings.get("settings_id"):
        settings_record = db.query(models.PaymentDocumentSettings).filter(
            models.PaymentDocumentSettings.settings_id == settings["settings_id"]
        ).first()
    if settings_record and settings_record.payment_authorization_enabled:
        blockers.append("Payment settings payment_authorization_enabled must be false.")
    if settings_record and settings_record.final_export_enabled:
        blockers.append("Payment settings final_export_enabled must be false.")
    live_count = sum(len(schedule.supervisions) for schedule in source["schedules"])
    if live_count == 0:
        blockers.append("No live post-optimize Supervision rows exist for the requested scope.")
    unresolved_names = sum(
        1
        for schedule in source["schedules"]
        for supervision in schedule.supervisions
        if not supervision.user or not (supervision.user.full_name or supervision.user.username)
    )
    if unresolved_names:
        blockers.append(f"{unresolved_names} live Supervision row(s) have no resolvable assigned-person name.")
    return {
        "available": not blockers,
        "review_status": review_status,
        "settings_status": settings.get("settings_status"),
        "settings_source_status": settings["settings_source_status"],
        "post_optimize_roster_status": "AVAILABLE" if live_count else "MISSING",
        "live_supervision_count": live_count,
        "paper_distribution_assignment_count": len(source["paper_assignments"]),
        "paper_distribution_mapping_status": "AVAILABLE" if source["paper_assignments"] else "MISSING_OR_EMPTY",
        "blocked_reasons": list(dict.fromkeys(blockers)),
        "payment_authorization_enabled": False,
        "final_export_enabled": False,
        "export_status": EXPORT_STATUS,
    }


def get_finance_support_roster_status(db: Session, request_payload: dict[str, Any]) -> dict[str, Any]:
    return _gate_status(db, load_supporting_roster_sources(db, request_payload))


def _banner(ws, ncols: int) -> None:
    yellow = PatternFill("solid", fgColor="FFF2CC")
    for row, text in ((1, DRAFT_LABEL_TH), (2, DRAFT_LABEL_EN), (3, DRAFT_STATUS)):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row, 1, text)
        cell.font = Font(bold=True, size=12 if row == 1 else 10)
        cell.fill = yellow
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_sheet(ws, headers: list[str], rows: list[list[Any]], widths: list[int], money_cols: set[int] | None = None) -> None:
    _banner(ws, len(headers))
    header_row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, values in enumerate(rows, header_row + 1):
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_index, col, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if money_cols and col in money_cols and value is not None:
                cell.number_format = '#,##0.00'
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _build_workbook(source: dict[str, Any], data: dict[str, Any]) -> "openpyxl.Workbook":
    if not _OPENPYXL_AVAILABLE:
        raise HTTPException(500, "openpyxl is not installed; xlsx export unavailable.")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    summary_rows = [[r[k] for k in ("date", "weekday", "day_type", "time", "physical_rooms", "course_sections", "students", "inv_count", "paper_count", "total_count", "rate", "inv_amount", "paper_amount", "total_amount", "note")] for r in data["summaries"]]
    person_rows = [[r["date"], r["day_type"], r["time"], r["name"], r["role_label"], r["rooms_text"], r["refs_text"], r["payment_count"], r["rate"], r["amount"], r["sources_text"], r["status_text"], r["note"], ""] for r in data["people"]]
    signature_rows = [[index, r["date"], r["time"], r["name"], r["role_label"], r["amount"], "", r["status_text"]] for index, r in enumerate(data["people"], 1)]
    mapping_rows = [[r[k] for k in ("date", "time", "rank", "room", "students", "refs", "person", "reason", "status", "note")] for r in data["mappings"]]
    trace_rows = [[r[k] for k in ("date", "time", "course", "section", "instructor", "students", "room", "room_type", "supervision_names", "paper_names", "schedule_id", "note")] for r in data["traces"]]
    specs = [
        (SHEET_NAMES[0], ["วันที่สอบ", "วัน", "ประเภทวัน", "ช่วงเวลา", "ห้องจริง", "วิชา/ตอน", "นักศึกษา", "คุมสอบ/เปิดห้อง", "จ่ายข้อสอบที่ผูกห้อง", "รวมคน", "อัตรา", "เงินคุมสอบ/เปิดห้อง", "เงินจ่ายข้อสอบ", "รวมเงิน", "หมายเหตุ"], summary_rows, [14, 12, 12, 16, 10, 12, 12, 16, 18, 12, 12, 18, 16, 16, 32], {11, 12, 13, 14}),
        (SHEET_NAMES[1], ["วันที่สอบ", "ประเภทวัน", "เวลา", "ชื่อ-นามสกุล", "บทบาท", "ห้องที่เกี่ยวข้อง", "วิชา/ตอน", "จำนวนครั้ง", "อัตรา", "จำนวนเงิน", "แหล่งที่มา", "สถานะการแมป", "หมายเหตุ", "ตรวจลายเซ็น"], person_rows, [14, 12, 16, 24, 24, 22, 22, 12, 12, 14, 28, 35, 34, 16], {9, 10}),
        (SHEET_NAMES[2], ["ลำดับ", "วันที่สอบ", "เวลา", "ชื่อ-นามสกุล", "บทบาท", "จำนวนเงิน", "ลายเซ็นผู้รับเงิน", "หมายเหตุ"], signature_rows, [8, 14, 16, 26, 24, 14, 28, 36], {6}),
        (SHEET_NAMES[3], ["วันที่สอบ", "เวลา", "ลำดับห้อง", "ห้อง", "นักศึกษา", "วิชา/ตอน", "ผู้จ่ายข้อสอบ/เปิดห้อง", "เหตุผลการผูก", "สถานะการแมป", "หมายเหตุ"], mapping_rows, [14, 16, 12, 16, 12, 24, 26, 26, 38, 34], set()),
        (SHEET_NAMES[4], ["วันที่สอบ", "เวลา", "วิชา", "ตอน", "ผู้สอน (อ้างอิงเท่านั้น)", "นักศึกษา", "ห้อง", "ประเภทห้อง", "ชื่อคุมสอบจาก Supervision", "ชื่อจ่ายข้อสอบจาก PDA", "source schedule id", "หมายเหตุ"], trace_rows, [14, 16, 14, 10, 26, 12, 18, 14, 34, 34, 18, 30], set()),
    ]
    for name, headers, rows, widths, money_cols in specs:
        _write_sheet(wb.create_sheet(name), headers, rows, widths, money_cols)
    summary_ws = wb[SHEET_NAMES[0]]
    summary_total_row = 6 + len(summary_rows)
    summary_ws.cell(summary_total_row, 1, "รวม")
    for column in (5, 6, 7, 8, 9, 10, 12, 13, 14):
        summary_ws.cell(summary_total_row, column, f"=SUM({get_column_letter(column)}6:{get_column_letter(column)}{summary_total_row - 1})")
        summary_ws.cell(summary_total_row, column).number_format = '#,##0.00' if column >= 12 else '#,##0'
    for cell in summary_ws[summary_total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    people_ws = wb[SHEET_NAMES[1]]
    people_total_row = 6 + len(person_rows)
    people_ws.cell(people_total_row, 1, "รวม")
    people_ws.cell(people_total_row, 8, f"=SUM(H6:H{people_total_row - 1})")
    people_ws.cell(people_total_row, 10, f"=SUM(J6:J{people_total_row - 1})")
    people_ws.cell(people_total_row, 10).number_format = '#,##0.00'
    for cell in people_ws[people_total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    signature_ws = wb[SHEET_NAMES[2]]
    previous_group = None
    group_fill = PatternFill("solid", fgColor="F2F7FB")
    use_fill = False
    for row_index in range(6, 6 + len(signature_rows)):
        group = (signature_ws.cell(row_index, 2).value, signature_ws.cell(row_index, 3).value)
        if group != previous_group:
            use_fill = not use_fill
            previous_group = group
        if use_fill:
            for cell in signature_ws[row_index]:
                cell.fill = group_fill
    ws = wb[SHEET_NAMES[1]]
    metadata = [
        ("export_status", EXPORT_STATUS), ("payment_authorization_enabled", "false"), ("final_export_enabled", "false"),
        ("source", "post-optimize Supervision"), ("paper_source", "PaperDistributionAssignment"),
        ("term", source["term"]["term_label"]), ("exam_type", source["term"]["exam_type"]),
        ("generated_at", datetime.now(timezone.utc).isoformat()),
    ]
    start = 8 + len(person_rows)
    for offset, (key, value) in enumerate(metadata):
        ws.cell(start + offset, 1, key)
        ws.cell(start + offset, 2, value)
    apply_workbook_thai_style(wb, title_rows=(1,), header_rows=(2, 3, 5))
    return wb


def build_finance_support_roster_export(db: Session, request_payload: dict[str, Any]) -> tuple["openpyxl.Workbook", str]:
    source = load_supporting_roster_sources(db, request_payload)
    status = _gate_status(db, source)
    if not status["available"]:
        raise HTTPException(400, "Supporting roster export gate failed: " + "; ".join(status["blocked_reasons"]))
    data = _build_data(source)
    wb = _build_workbook(source, data)
    generated_at = datetime.now(timezone.utc)
    filename = f"EMS_DRAFT_FINANCE_ROSTER_{source['term']['semester']}-{source['term']['academic_year']}_{generated_at.strftime('%Y%m%d_%H%M')}.xlsx"
    return wb, filename
