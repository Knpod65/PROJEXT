"""
export_excel_service.py — Excel export orchestration.

Owns:
- workbook generation orchestration
- filename generation
- sheet metadata shaping
- audit payload shaping
"""
from typing import Optional, Tuple
from datetime import date
from sqlalchemy.orm import Session
import models
from repositories.export_repository import ExportRepository
from serializers.export_serializer import serialize_export_metadata


class ExportExcelService:
    """Orchestration for Excel export operations."""

    @staticmethod
    def get_compensation_workbook(db: Session, data: dict, exam_type: str) -> Tuple[object, str]:
        """Build compensation workbook."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "ค่าตอบแทนรายบุคคล"
        ws1.row_dimensions[1].height = 30

        headers1 = ["ลำดับ", "รหัสพนักงาน", "ชื่อ-สกุล", "ตำแหน่ง", "สังกัด", "จำนวนครั้งคุมสอบ", "ค่าตอบแทน/ครั้ง (บาท)", "รวม (บาท)"]
        for c, h in enumerate(headers1, 1):
            ws1.cell(1, c, h)

        user_data = {}
        sups = data["supervisions"]
        rate_internal = data["rate_internal"]
        rate_external = data["rate_external"]

        for sup in sups:
            if not sup.user:
                continue
            uid = sup.user_id
            if uid not in user_data:
                user_data[uid] = {"user": sup.user, "count": 0, "rate": rate_internal, "total": 0.0}
            user_data[uid]["count"] += 1
            rate = float(sup.compensation) if sup.compensation else rate_internal
            user_data[uid]["rate"] = rate
            user_data[uid]["total"] += rate

        for i, (uid, d) in enumerate(sorted(user_data.items(), key=lambda x: x[1]["user"].full_name or ""), 1):
            u = d["user"]
            row = [i, u.employee_id or "", u.full_name or u.username, u.title or "", u.division or "", d["count"], d["rate"], d["total"]]
            ws1.append(row)
            if i % 2 == 0:
                for c in range(1, len(headers1) + 1):
                    ws1.cell(i + 1, c).fill = PatternFill("solid", fgColor="F8FAFC")
            ws1.cell(i + 1, 8).font = Font(bold=True)
            ws1.cell(i + 1, 8).number_format = '#,##0.00'

        total_row = len(user_data) + 2
        ws1.cell(total_row, 6, sum(d["count"] for d in user_data.values()))
        ws1.cell(total_row, 8, sum(d["total"] for d in user_data.values()))
        ws1.cell(total_row, 6).font = Font(bold=True, color="C41230")
        ws1.cell(total_row, 8).font = Font(bold=True, color="C41230")
        ws1.cell(total_row, 8).number_format = '#,##0.00'
        ws1.cell(total_row, 5, "รวมทั้งสิ้น").font = Font(bold=True)

        ws2 = wb.create_sheet("รายวิชา")
        headers2 = ["รหัสวิชา", "ชื่อวิชา", "ตอน", "วันสอบ", "เวลา", "ห้อง", "จำนวน นศ.", "กรรมการ 1", "กรรมการ 2", "กรรมการ 3", "รวมค่าตอบแทน"]
        for c, h in enumerate(headers2, 1):
            ws2.cell(1, c, h)

        schedules = {}
        for sup in sups:
            sid = sup.schedule_id
            if sid not in schedules:
                schedules[sid] = {"sch": sup.schedule, "sups": []}
            schedules[sid]["sups"].append(sup)

        for i, (sid, d) in enumerate(sorted(schedules.items(), key=lambda x: (str(x[1]["sch"].exam_date or ""), x[1]["sch"].exam_time or "")), 1):
            sch = d["sch"]
            sec = sch.section
            course = sec.course if sec else None
            room = sch.room
            sups_sorted = sorted(d["sups"], key=lambda s: s.slot_order or 99)
            sup_names = [s.user.full_name if s.user else "" for s in sups_sorted[:3]]
            while len(sup_names) < 3:
                sup_names.append("")
            total_comp = sum(float(s.compensation or rate_internal) for s in d["sups"])
            row = [course.course_id if course else "", course.course_name_th if course else "", sec.section_no if sec else "", str(sch.exam_date) if sch.exam_date else "", sch.exam_time or "", room.room_name if room else "", sec.num_students if sec else 0, sup_names[0], sup_names[1], sup_names[2], total_comp]
            ws2.append(row)
            ws2.cell(i + 1, 11).number_format = '#,##0.00'
            if i % 2 == 0:
                for c in range(1, len(headers2) + 1):
                    ws2.cell(i + 1, c).fill = PatternFill("solid", fgColor="F8FAFC")

        ws3 = wb.create_sheet("สรุป")
        exam_type_th = "ปลายภาค" if exam_type == "final" else "กลางภาค"
        info = [["รายงานค่าตอบแทนกรรมการคุมสอบ", ""], ["ภาคการศึกษา", f"{data.get('semester', '')}/{data.get('academic_year', '')}"], ["ประเภทสอบ", exam_type_th], ["วันที่ออกรายงาน", str(date.today())], ["", ""], ["จำนวนวิชาที่สอบ", len(schedules)], ["จำนวนกรรมการทั้งหมด", len(user_data)], ["จำนวนครั้งคุมสอบรวม", sum(d["count"] for d in user_data.values())], ["ค่าตอบแทนรวมทั้งสิ้น", sum(d["total"] for d in user_data.values())]]
        for r, (k, v) in enumerate(info, 1):
            ws3.cell(r, 1, k).font = Font(bold=(r == 1 or r >= 6))
            ws3.cell(r, 2, v)
            if k == "ค่าตอบแทนรวมทั้งสิ้น":
                ws3.cell(r, 2).number_format = '#,##0.00'
                ws3.cell(r, 2).font = Font(bold=True, color="C41230", size=13)

        filename = f"EMS_compensation_{data.get('semester', '')}_{data.get('academic_year', '')}_{exam_type}.xlsx"
        return wb, filename

    @staticmethod
    def get_schedule_workbook(db: Session, data: dict) -> Tuple[object, str]:
        """Build schedule workbook."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ตารางสอบ"
        headers = ["วันที่", "เวลา", "รหัสวิชา", "ชื่อวิชา", "ตอน", "ห้อง", "จำนวน นศ.", "อาจารย์", "กรรมการ 1", "กรรมการ 2", "กรรมการ 3", "สถานะ"]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)

        for i, s in enumerate(data.get("schedules", []), 1):
            sec = s.section
            course = sec.course if sec else None
            room = s.room
            teacher = sec.teacher if sec else None
            sups = sorted(s.supervisions or [], key=lambda x: x.slot_order or 99)
            names = [sv.user.full_name for sv in sups[:3] if sv.user]
            while len(names) < 3:
                names.append("")
            row = [str(s.exam_date) if s.exam_date else "", s.exam_time or "", course.course_id if course else "", course.course_name_th if course else "", sec.section_no if sec else "", room.room_name if room else "", sec.num_students if sec else 0, teacher.full_name if teacher else "", names[0], names[1], names[2], s.status.value if s.status else ""]
            ws.append(row)
            if i % 2 == 0:
                for c in range(1, len(headers)+1):
                    ws.cell(i+1, c).fill = PatternFill("solid", fgColor="F8FAFC")

        filename = f"EMS_schedule_{data.get('semester', '')}_{data.get('academic_year', '')}_{data.get('exam_type', '')}.xlsx"
        return wb, filename

    @staticmethod
    def get_submissions_workbook(db: Session, data: dict) -> Tuple[object, str]:
        """Build submissions workbook."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        status_th = {
            "draft": "ร่าง", "submitted": "รอตรวจ", "approved": "อนุมัติ",
            "rejected": "ปฏิเสธ", "released": "ปล่อยแล้ว",
        }
        STATUS_COLORS = {
            "draft": "D4D4D4", "submitted": "FEF3C7", "approved": "DCFCE7",
            "rejected": "FEE2E2", "released": "DBEAFE",
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "สถานะข้อสอบ"

        headers = [
            "รหัสวิชา", "ชื่อวิชา", "ตอน", "อาจารย์ผู้รับผิดชอบ",
            "รูปแบบสอบ", "อัปโหลด PDF", "สเปคพิมพ์", "สถานะ",
            "กระดาษเขียนตอบ", "สมุดคำตอบ", "OMR", "หมายเหตุ",
            "อัปเดตล่าสุด",
        ]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)

        for i, s in enumerate(data.get("submissions", []), 1):
            sec = s.section
            course = sec.course if sec else None
            mat = s.material_request
            row = [
                course.course_id if course else "",
                course.course_name_th if course else "",
                sec.section_no if sec else "",
                s.submitter.full_name if s.submitter else "",
                s.exam_type_choice or "—",
                "✅" if s.has_uploaded_pdf else "❌",
                "✅" if s.print_spec_confirmed else "❌",
                status_th.get(s.status.value if s.status else "", "—"),
                mat.answer_paper_sheets if mat else 0,
                mat.answer_booklet_count if mat else 0,
                mat.omr_sheet_count if mat else 0,
                mat.special_note if mat else "",
                str(s.submitted_at.date()) if s.submitted_at else "",
            ]
            ws.append(row)
            color = STATUS_COLORS.get(s.status.value if s.status else "draft", "FFFFFF")
            ws.cell(i+1, 8).fill = PatternFill("solid", fgColor=color)
            ws.cell(i+1, 8).font = Font(bold=True)

        filename = f"EMS_submissions_{data.get('semester', '')}_{data.get('academic_year', '')}.xlsx"
        return wb, filename

    @staticmethod
    def get_workload_summary_workbook(db: Session, data: dict) -> Tuple[object, str]:
        """Build workload summary workbook."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workload Summary"
        headers = [
            "Staff Name", "Department", "Invigilation", "Paper Distribution",
            "External Exam", "Current Total", "Historical Total",
        ]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)

        for row in data.get("summary", []):
            ws.append([
                row["staff_name"], row["department"],
                row["invigilation_count"], row["paper_distribution_count"],
                row["external_exam_count"], row["total_workload"],
                row["historical_total_workload"],
            ])

        period = data.get("period", {})
        filename = f"EMS_workload_summary_{period.get('semester', '')}_{period.get('academic_year', '')}_{period.get('exam_type', '')}.xlsx"
        return wb, filename

    @staticmethod
    def get_workload_detail_workbook(db: Session, data: dict) -> Tuple[object, str]:
        """Build workload detail workbook."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workload Detail"
        headers = [
            "Staff Name", "Department", "Duty Type", "Date", "Time",
            "Context", "Room", "Workload Count",
        ]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)

        for row in data.get("details", []):
            ws.append([
                row["staff_name"], row["department"], row["duty_type"],
                row["date"], row["time"], row["context_label"],
                row["room"], row["workload_count"],
            ])

        period = data.get("period", {})
        filename = f"EMS_workload_detail_{period.get('semester', '')}_{period.get('academic_year', '')}_{period.get('exam_type', '')}.xlsx"
        return wb, filename

    @staticmethod
    def get_paper_distribution_workbook(db: Session, data: dict) -> Tuple[object, str]:
        """Build paper distribution workbook."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Paper Distribution"
        headers = [
            "Staff Name", "Department", "Date", "Time",
            "Covered Courses", "Covered Rooms", "Schedules Covered",
            "Workload Count", "Assignment Mode",
        ]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)

        rows = data.get("rows", [])
        slot_context = data.get("slot_context", {})
        for row in rows:
            context = slot_context.get((row.exam_date, row.exam_time), {"courses": [], "rooms": []})
            ws.append([
                row.user.full_name if row.user else f"User #{row.user_id}",
                (row.user.division or row.user.unit) if row.user else "",
                row.exam_date, row.exam_time,
                ", ".join(context["courses"]), ", ".join(context["rooms"]),
                row.covered_schedule_count or 0, row.workload_units or 1,
                row.assignment_mode,
            ])

        period = data.get("period", {})
        filename = f"EMS_paper_distribution_{period.get('semester', '')}_{period.get('academic_year', '')}_{period.get('exam_type', '')}.xlsx"
        return wb, filename