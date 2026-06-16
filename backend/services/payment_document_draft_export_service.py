"""Draft-only payment document export service.

Generates a gated xlsx workbook from the existing draft preview.
No DB writes. Safety flags (payment_authorization_enabled,
final_export_enabled) are never set to true.
"""
from __future__ import annotations

from datetime import datetime, timezone
from typing import Any

from fastapi import HTTPException
from sqlalchemy.orm import Session

import models
from services.official_payment_document_draft_service import (
    CALCULATED_FROM_SETTINGS,
    build_official_payment_document_draft_preview,
)
from services.payment_document_settings_service import SETTINGS_CONFIGURED
from services.thai_export_service import apply_workbook_thai_style

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


ACCEPTED_STATUS = "ACCEPTED_FOR_DRAFT_EXPORT"
ACTIVE_SETTINGS_STATUS = "ACTIVE_FOR_DRAFT_PREVIEW"
DRAFT_LABEL_TH = "ร่างเอกสารเพื่อการตรวจทานเท่านั้น ยังไม่ใช่เอกสารอนุมัติเบิกจ่าย"
DRAFT_LABEL_EN = "Draft for review only. Not payment authorization."
DRAFT_FOOTER_TH = "เอกสารนี้เป็นร่างเพื่อการตรวจทาน ไม่ใช่เอกสารอนุมัติเบิกจ่าย"


def _document_id(payload: dict[str, Any]) -> str:
    ay = payload.get("academic_year") or "unknown"
    sem = payload.get("semester") or "unknown"
    et = payload.get("exam_type") or "unknown"
    pid = payload.get("period_id")
    return f"ADVANCE_PAYMENT_DRAFT_SUMMARY:{ay}:{sem}:{et}:{pid if pid is not None else 'all'}"


def _find_accepted_review(
    db: Session, document_id: str
) -> "models.PaymentDocumentReviewRecord":
    record = (
        db.query(models.PaymentDocumentReviewRecord)
        .filter(
            models.PaymentDocumentReviewRecord.document_id == document_id,
            models.PaymentDocumentReviewRecord.review_status == ACCEPTED_STATUS,
        )
        .order_by(models.PaymentDocumentReviewRecord.id.desc())
        .first()
    )
    if record is None:
        raise HTTPException(
            400,
            f"No {ACCEPTED_STATUS} review record found for document '{document_id}'. "
            "Export requires an accepted review.",
        )
    if not (record.comment or "").strip():
        raise HTTPException(
            400,
            "Review record comment is required for export but is empty.",
        )
    return record


def _check_gate(draft: dict[str, Any]) -> None:
    meta = draft.get("metadata") or {}
    checks = [
        (
            meta.get("settings_source_status") != SETTINGS_CONFIGURED,
            "settings_source_status must be CONFIGURED",
        ),
        (
            meta.get("settings_status") != ACTIVE_SETTINGS_STATUS,
            "settings_status must be ACTIVE_FOR_DRAFT_PREVIEW",
        ),
        (
            meta.get("calculation_status") != CALCULATED_FROM_SETTINGS,
            "calculation_status must be CALCULATED_FROM_SETTINGS",
        ),
        (
            not (meta.get("paper_distribution_responsible_group") or "").strip(),
            "paper_distribution_responsible_group must be set in settings",
        ),
        (
            draft.get("payment_authorization_enabled"),
            "payment_authorization_enabled must be false",
        ),
        (
            draft.get("final_export_enabled"),
            "final_export_enabled must be false",
        ),
    ]
    for failed, message in checks:
        if failed:
            raise HTTPException(400, f"Draft export gate failed: {message}")


def _yellow_fill() -> "PatternFill":
    return PatternFill("solid", fgColor="FFFF00")


def _build_workbook(
    draft: dict[str, Any],
    review: "models.PaymentDocumentReviewRecord",
    generated_at: datetime,
) -> "openpyxl.Workbook":
    if not _OPENPYXL_AVAILABLE:
        raise HTTPException(500, "openpyxl is not installed; xlsx export unavailable.")

    meta = draft.get("metadata") or {}
    rows = draft.get("rows") or []
    totals = draft.get("totals") or {}

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "ร่างเอกสาร"

    NCOLS = 9
    col_range = f"A1:{get_column_letter(NCOLS)}"

    # --- Draft banners (rows 1-3) ---
    ws1.merge_cells(f"{col_range}1")
    ws1["A1"] = DRAFT_LABEL_TH
    ws1["A1"].font = Font(bold=True, size=13)
    ws1["A1"].fill = _yellow_fill()
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1.merge_cells(f"A2:{get_column_letter(NCOLS)}2")
    ws1["A2"] = DRAFT_LABEL_EN
    ws1["A2"].font = Font(bold=True)
    ws1["A2"].fill = _yellow_fill()
    ws1["A2"].alignment = Alignment(horizontal="center")

    ws1.merge_cells(f"A3:{get_column_letter(NCOLS)}3")
    ws1["A3"] = "DRAFT_NOT_AUTHORIZED"
    ws1["A3"].fill = _yellow_fill()
    ws1["A3"].alignment = Alignment(horizontal="center")

    # --- Title and metadata (rows 5-8) ---
    ws1.merge_cells(f"A5:{get_column_letter(NCOLS)}5")
    ws1["A5"] = "สรุปจำนวนกรรมการและค่าตอบแทน รายวัน/ช่วงเวลา"
    ws1["A5"].font = Font(bold=True, size=12)
    ws1["A5"].alignment = Alignment(horizontal="center")

    term_label = meta.get("term_label") or f"{meta.get('semester')}/{meta.get('academic_year')}"
    currency = meta.get("currency") or "THB"
    wd_rate = meta.get("weekday_rate")
    we_rate = meta.get("weekend_rate")

    ws1.merge_cells(f"A6:{get_column_letter(NCOLS)}6")
    ws1["A6"] = f"ภาคการศึกษา: {term_label}"

    ws1.merge_cells(f"A7:{get_column_letter(NCOLS)}7")
    ws1["A7"] = (
        f"อัตราวันธรรมดา: {wd_rate} {currency}  |  "
        f"อัตราวันหยุด: {we_rate} {currency}"
    )

    ws1.merge_cells(f"A8:{get_column_letter(NCOLS)}8")
    ws1["A8"] = (
        f"หน่วยงานรับผิดชอบจ่ายข้อสอบ: "
        f"{meta.get('paper_distribution_responsible_group', '')}"
    )

    # --- Column headers (row 9) ---
    headers = [
        "วันที่สอบ",
        "ช่วงเวลา",
        "ประเภทวัน",
        "จำนวนกรรมการคุมสอบ",
        "ค่าตอบแทนคุมสอบ",
        "จำนวนกรรมการจ่ายข้อสอบ",
        "ค่าตอบแทนจ่ายข้อสอบ",
        "รวมค่าตอบแทน",
        "อัตรา",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws1.cell(row=9, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # --- Data rows (row 10+) ---
    def _float(value: Any) -> float | None:
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    for r, row in enumerate(rows, start=10):
        ws1.cell(row=r, column=1, value=row.get("exam_date") or row.get("normalized_exam_date"))
        ws1.cell(row=r, column=2, value=row.get("time_slot"))
        ws1.cell(row=r, column=3, value=row.get("day_type"))
        ws1.cell(row=r, column=4, value=row.get("invigilation_committee_count"))
        ws1.cell(row=r, column=5, value=_float(row.get("invigilation_compensation_amount")))
        ws1.cell(row=r, column=6, value=row.get("paper_distribution_committee_count"))
        ws1.cell(row=r, column=7, value=_float(row.get("paper_distribution_compensation_amount")))
        ws1.cell(row=r, column=8, value=_float(row.get("total_compensation_amount")))
        ws1.cell(row=r, column=9, value=_float(row.get("rate_amount")))

    # --- Totals row ---
    totals_row = 10 + len(rows)
    ws1.cell(row=totals_row, column=1, value="รวม").font = Font(bold=True)
    ws1.cell(row=totals_row, column=4, value=totals.get("invigilation_committee_count"))
    ws1.cell(row=totals_row, column=5, value=_float(totals.get("invigilation_compensation_amount")))
    ws1.cell(row=totals_row, column=6, value=totals.get("paper_distribution_committee_count"))
    ws1.cell(row=totals_row, column=7, value=_float(totals.get("paper_distribution_compensation_amount")))
    ws1.cell(row=totals_row, column=8, value=_float(totals.get("grand_total_amount")))
    for col_idx in range(1, NCOLS + 1):
        ws1.cell(row=totals_row, column=col_idx).font = Font(bold=True)

    # --- Footer draft label ---
    footer_row = totals_row + 2
    ws1.merge_cells(f"A{footer_row}:{get_column_letter(NCOLS)}{footer_row}")
    ws1.cell(row=footer_row, column=1, value=DRAFT_FOOTER_TH)
    ws1.cell(row=footer_row, column=1).fill = _yellow_fill()
    ws1.cell(row=footer_row, column=1).alignment = Alignment(horizontal="center")

    # --- Column widths (use get_column_letter to avoid MergedCell AttributeError) ---
    col_widths = [20, 18, 12, 22, 22, 24, 24, 18, 12]
    for col_idx, width in enumerate(col_widths, start=1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Sheet 2: review metadata ---
    ws2 = wb.create_sheet("การตรวจร่าง")
    reviewed_at_str = str(review.reviewed_at) if review.reviewed_at else ""
    review_rows = [
        ("ผู้ตรวจ (Reviewer)", review.reviewer_name or ""),
        ("บทบาท (Role)", review.reviewer_role or ""),
        ("สถานะการตรวจ (Review Status)", review.review_status or ""),
        ("วันที่ตรวจ (Reviewed At)", reviewed_at_str),
        ("หมายเหตุ (Comment)", review.comment or ""),
        ("", ""),
        ("แหล่งที่มาของการตั้งค่า (Settings Source)", meta.get("rate_source", "")),
        ("สถานะการตั้งค่า (Settings Status)", meta.get("settings_status", "")),
        ("สถานะการคำนวณ (Calculation Status)", meta.get("calculation_status", "")),
        ("", ""),
        ("สถานะเอกสาร (Document Status)", "DRAFT_NOT_AUTHORIZED"),
        ("payment_authorization_enabled", "false"),
        ("final_export_enabled", "false"),
        ("", ""),
        ("สร้างเมื่อ (Generated At)", generated_at.isoformat()),
    ]
    for row_idx, (key, val) in enumerate(review_rows, start=1):
        ws2.cell(row=row_idx, column=1, value=key)
        ws2.cell(row=row_idx, column=2, value=val)

    ws2.column_dimensions[get_column_letter(1)].width = 42
    ws2.column_dimensions[get_column_letter(2)].width = 60

    apply_workbook_thai_style(wb, title_rows=(1, 5), header_rows=(2, 3, 9))
    return wb


def _export_filename(payload: dict[str, Any], generated_at: datetime) -> str:
    sem = payload.get("semester") or "X"
    ay = payload.get("academic_year") or "XXXX"
    ts = generated_at.strftime("%Y%m%d_%H%M")
    return f"EMS_DRAFT_PAYMENT_DOCUMENT_{sem}-{ay}_{ts}.xlsx"


def build_payment_document_draft_export(
    db: Session, request_payload: dict[str, Any]
) -> tuple["openpyxl.Workbook", str]:
    """Gate-checked draft export. Returns (workbook, filename).

    Raises HTTPException 400 if any precondition fails.
    Never sets payment_authorization_enabled or final_export_enabled to true.
    """
    document_id = _document_id(request_payload)
    review = _find_accepted_review(db, document_id)
    draft = build_official_payment_document_draft_preview(db, request_payload)
    _check_gate(draft)
    generated_at = datetime.now(timezone.utc)
    wb = _build_workbook(draft, review, generated_at)
    filename = _export_filename(request_payload, generated_at)
    return wb, filename
