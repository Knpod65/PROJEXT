"""
exports_excel.py — Export Excel รายงาน

Thin router that delegates to:
- ExportExcelService for workbook generation
- ExportService for data access
- ExportValidator for validation
- ExportPolicy for permissions
"""
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Request
from sqlalchemy.orm import Session
from database import get_db
import models
from auth_utils import require_admin, require_staff_or_admin, get_current_user, require_view_all, log_action
from config.audit_actions import (
    EXPORT_COMPENSATION,
    EXPORT_PAPER_DISTRIBUTION_EXCEL,
    EXPORT_SCHEDULE_EXCEL,
    EXPORT_SUBMISSIONS_EXCEL,
    EXPORT_WORKLOAD_DETAIL_EXCEL,
    EXPORT_WORKLOAD_SUMMARY_EXCEL,
)
from config.periods import resolve_export_period
from services.export_service import ExportService
from services.export_excel_service import ExportExcelService
from services.thai_export_service import workbook_streaming_response
from validators.export_validator import ExportValidator
from validators.export_excel_validator import ExportExcelValidator
from policies.export_policy import ExportPolicy

router = APIRouter()


def _workbook_response(wb, filename: str):
    return workbook_streaming_response(wb, filename)


def _style_header(ws, row: int, cols: int, fill_color="0F1B35", font_color="FFFFFF"):
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        fill = PatternFill("solid", fgColor=fill_color)
        font = Font(bold=True, color=font_color, size=11)
        bd = Border(bottom=Side(style="medium", color="C41230"))
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bd
    except ImportError:
        pass


def _auto_width(ws):
    try:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value or "")
                    max_len = max(max_len, len(val))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
    except:
        pass


@router.get("/compensation")
def export_compensation(
    semester: str = None,
    academic_year: str = None,
    exam_type: str = "final",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
    request: Request = None,
):
    ExportValidator.validate_exam_type(exam_type)
    ExportValidator.validate_semester(semester)
    ExportPolicy.require_export_permission(current_user, "compensation")

    if not semester or not academic_year:
        p = db.query(models.ExamPeriod).filter(models.ExamPeriod.is_active == True).first()
        if p:
            semester = semester or p.semester
            academic_year = academic_year or p.academic_year

    data = ExportService.get_compensation_data(db, semester, academic_year, exam_type)
    wb, filename = ExportExcelService.get_compensation_workbook(db, data, exam_type)

    try:
        log_action(
            db=db, actor=current_user, action=EXPORT_COMPENSATION,
            table_name="supervisions",
            new_values={"file_type": "xlsx", "export_scope": "compensation", "row_count": len(data.get("supervisions", [])), "semester": semester, "academic_year": academic_year, "exam_type": exam_type},
            request=request, http_status=200
        )
    except Exception:
        pass

    return _workbook_response(wb, filename)


@router.get("/schedule-excel")
def export_schedule_excel(
    semester: str = None,
    academic_year: str = None,
    exam_type: str = "final",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_view_all),
    request: Request = None,
):
    ExportValidator.validate_exam_type(exam_type)
    ExportValidator.validate_semester(semester)
    ExportPolicy.require_export_permission(current_user, "schedule")

    if not semester or not academic_year:
        p = db.query(models.ExamPeriod).filter(models.ExamPeriod.is_active == True).first()
        if p:
            semester = semester or p.semester
            academic_year = academic_year or p.academic_year

    scheds = ExportService.get_schedule_export_data(db, semester, academic_year, exam_type)
    data = {"schedules": scheds, "semester": semester, "academic_year": academic_year, "exam_type": exam_type}
    wb, filename = ExportExcelService.get_schedule_workbook(db, data)

    try:
        log_action(
            db=db, actor=current_user, action=EXPORT_SCHEDULE_EXCEL,
            table_name="exam_schedules",
            new_values={"file_type": "xlsx", "export_scope": "schedule", "row_count": len(scheds), "semester": semester, "academic_year": academic_year, "exam_type": exam_type},
            request=request, http_status=200
        )
    except Exception:
        pass

    return _workbook_response(wb, filename)


@router.get("/submissions-excel")
def export_submissions_excel(
    semester: str = None,
    academic_year: str = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
    request: Request = None,
):
    ExportValidator.validate_semester(semester)
    ExportPolicy.require_export_permission(current_user, "submissions")

    if not semester or not academic_year:
        p = db.query(models.ExamPeriod).filter(models.ExamPeriod.is_active == True).first()
        if p:
            semester = semester or p.semester
            academic_year = academic_year or p.academic_year

    subs = ExportService.get_submissions_data(db, semester, academic_year)
    data = {"submissions": subs, "semester": semester, "academic_year": academic_year}
    wb, filename = ExportExcelService.get_submissions_workbook(db, data)

    try:
        log_action(
            db=db, actor=current_user, action=EXPORT_SUBMISSIONS_EXCEL,
            table_name="exam_submissions",
            new_values={"file_type": "xlsx", "export_scope": "submissions", "row_count": len(subs), "semester": semester, "academic_year": academic_year},
            request=request, http_status=200
        )
    except Exception:
        pass

    return _workbook_response(wb, filename)


@router.get("/workload-summary-excel")
def export_workload_summary_excel(
    semester: str = None,
    academic_year: str = None,
    exam_type: str = "final",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_staff_or_admin),
    request: Request = None,
):
    ExportValidator.validate_exam_type(exam_type)
    ExportPolicy.require_export_permission(current_user, "workload")

    period = resolve_export_period(db, semester, academic_year, exam_type)
    snapshot = ExportService.get_period_workload_snapshot(db, period)
    data = {"summary": snapshot["summary"], "period": {"semester": period.semester, "academic_year": period.academic_year, "exam_type": period.exam_type}}
    wb, filename = ExportExcelService.get_workload_summary_workbook(db, data)

    try:
        log_action(
            db=db, actor=current_user, action=EXPORT_WORKLOAD_SUMMARY_EXCEL,
            table_name="staffworkloads",
            new_values={"file_type": "xlsx", "export_scope": "workload_summary", "row_count": len(snapshot["summary"]), "semester": period.semester, "academic_year": period.academic_year, "exam_type": period.exam_type},
            request=request, http_status=200
        )
    except Exception:
        pass

    return _workbook_response(wb, filename)


@router.get("/workload-detail-excel")
def export_workload_detail_excel(
    semester: str = None,
    academic_year: str = None,
    exam_type: str = "final",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_staff_or_admin),
    request: Request = None,
):
    ExportValidator.validate_exam_type(exam_type)
    ExportPolicy.require_export_permission(current_user, "workload")

    period = resolve_export_period(db, semester, academic_year, exam_type)
    snapshot = ExportService.get_period_workload_snapshot(db, period)
    data = {"details": snapshot["details"], "period": {"semester": period.semester, "academic_year": period.academic_year, "exam_type": period.exam_type}}
    wb, filename = ExportExcelService.get_workload_detail_workbook(db, data)

    try:
        log_action(
            db=db, actor=current_user, action=EXPORT_WORKLOAD_DETAIL_EXCEL,
            table_name="staffworkloads",
            new_values={"file_type": "xlsx", "export_scope": "workload_detail", "row_count": len(snapshot["details"]), "semester": period.semester, "academic_year": period.academic_year, "exam_type": period.exam_type},
            request=request, http_status=200
        )
    except Exception:
        pass

    return _workbook_response(wb, filename)


@router.get("/paper-distribution-excel")
def export_paper_distribution_excel(
    semester: str = None,
    academic_year: str = None,
    exam_type: str = "final",
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_staff_or_admin),
    request: Request = None,
):
    ExportValidator.validate_exam_type(exam_type)
    ExportPolicy.require_export_permission(current_user, "paper_distribution")

    period = resolve_export_period(db, semester, academic_year, exam_type)
    assignments, schedules = ExportService.get_paper_distribution_data(db, period)

    slot_context = {}
    for schedule in schedules:
        key = (str(schedule.exam_date) if schedule.exam_date else "", schedule.exam_time or "")
        context = slot_context.setdefault(key, {"courses": [], "rooms": []})
        if schedule.section and schedule.section.course:
            label = f"{schedule.section.course.course_id} Sec {schedule.section.section_no}"
            if label not in context["courses"]:
                context["courses"].append(label)
        if schedule.room and schedule.room.room_name not in context["rooms"]:
            context["rooms"].append(schedule.room.room_name)

    data = {"rows": assignments, "slot_context": slot_context, "period": {"semester": period.semester, "academic_year": period.academic_year, "exam_type": period.exam_type}}
    wb, filename = ExportExcelService.get_paper_distribution_workbook(db, data)

    try:
        log_action(
            db=db, actor=current_user, action=EXPORT_PAPER_DISTRIBUTION_EXCEL,
            table_name="paper_distribution_assignments",
            new_values={"file_type": "xlsx", "export_scope": "paper_distribution", "row_count": len(assignments), "semester": period.semester, "academic_year": period.academic_year, "exam_type": period.exam_type},
            request=request, http_status=200
        )
    except Exception:
        pass

    return _workbook_response(wb, filename)
